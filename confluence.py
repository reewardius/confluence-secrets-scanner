# -*- coding: utf-8 -*-
import requests
from requests.auth import HTTPBasicAuth
import argparse
import csv
import re
from io import BytesIO
from PIL import Image
import pytesseract
import os
import warnings
import time
from datetime import datetime, timedelta
import logging
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

# Try to import YAML for TruffleHog format support
try:
    import yaml
except ImportError:
    yaml = None

try:
    import boto3
except ImportError:
    boto3 = None

os.environ["TESSDATA_PREFIX"] = "/usr/local/share/tessdata/"
warnings.simplefilter('ignore', Image.DecompressionBombWarning)

try:
    import docx
except ImportError:
    docx = None

try:
    import fitz
except ImportError:
    fitz = None

try:
    import zipfile
    import tarfile
except ImportError:
    zipfile = tarfile = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

def setup_logging(log_file="confluence_scan.log"):
    """Configure logging to console and file."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )

def check_dependencies():
    """Check for required dependencies and return any errors."""
    errors = []
    if not docx:
        errors.append("python-docx is required for .docx support")
    if not fitz:
        errors.append("PyMuPDF is required for .pdf support")
    if not zipfile or not tarfile:
        errors.append("zipfile and tarfile are required for archive support")
    if not Image or not pytesseract:
        errors.append("PIL and pytesseract are required for image OCR")
    if not Workbook:
        errors.append("openpyxl is required for .xlsx support (pip install openpyxl)")
    return errors

def safe_request(url, headers, auth=None, params=None):
    """Make a safe HTTP request with error handling."""
    try:
        resp = requests.get(url, headers=headers, auth=auth, params=params, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logging.error(f"Request {url}: {e}")
        return {}

def normalize_base_url(base_url):
    """Normalize base URL for Atlassian Cloud."""
    base_url = base_url.rstrip('/')
    if 'atlassian.net' in base_url and '/wiki' not in base_url:
        return f"{base_url}/wiki"
    return base_url

def download_attachment_with_redirect(url, auth, download_headers, timeout=30):
    """
    Download an attachment following Atlassian's 302 redirect to api.media.atlassian.com.

    Atlassian Cloud returns HTTP 302 from the REST download endpoint, redirecting to
    api.media.atlassian.com/file/{fileId}/binary?token=... The token in the redirect
    URL is already pre-authenticated, so we must NOT forward the Basic Auth header
    to the media domain (it would be ignored or cause issues). We use a Session with
    a response hook that strips Authorization on cross-domain redirects.
    """
    session = requests.Session()

    def strip_auth_on_redirect(r, *args, **kwargs):
        """Remove Authorization header when redirecting to a different host."""
        if r.is_redirect:
            redirect_url = r.headers.get("Location", "")
            original_host = requests.utils.urlparse(url).netloc
            redirect_host = requests.utils.urlparse(redirect_url).netloc
            if redirect_host and redirect_host != original_host:
                r.request.headers.pop("Authorization", None)

    session.hooks["response"].append(strip_auth_on_redirect)

    try:
        resp = session.get(
            url,
            auth=auth,
            headers=download_headers,
            timeout=timeout,
            allow_redirects=True,
        )
        resp.raise_for_status()
        return resp, None
    except requests.exceptions.HTTPError as e:
        status = e.response.status_code if e.response is not None else 0
        return None, status
    except Exception:
        return None, -1


def get_all_spaces(base_url, auth, headers):
    """Retrieve all spaces from Confluence."""
    all_spaces, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/space"
        data = safe_request(url, headers, auth, params={"start": start, "limit": limit})
        if not data:
            break
        all_spaces.extend(data.get("results", []))
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_spaces

def get_public_spaces(base_url, headers):
    """Retrieve only public spaces without authentication."""
    all_spaces, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/space"
        data = safe_request(url, headers, auth=None, params={"start": start, "limit": limit})
        if not data:
            break
        spaces = data.get("results", [])
        for space in spaces:
            space_key = space.get("key")
            test_url = f"{base_url}/rest/api/content"
            test_data = safe_request(
                test_url, headers, auth=None,
                params={"type": "page", "spaceKey": space_key, "limit": 1}
            )
            if test_data and test_data.get("results"):
                all_spaces.append(space)
                logging.info(f"Found public space: {space_key}")
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_spaces

def get_pages_in_space(base_url, auth, headers, space_key, modified_after=None, modified_before=None, created_in_years=None, modified_in_years=None):
    """Retrieve pages in a space, applying date filters.
    
    Args:
        created_in_years: None, single year (int), or list of years ([int])
    """
    all_pages, start, limit = [], 0, 50
    while True:
        url = f"{base_url}/rest/api/content"
        params = {"type": "page", "spaceKey": space_key, "start": start, "limit": limit, "expand": "history,version"}
        data = safe_request(url, headers, auth, params=params)
        if not data:
            break
        pages = data.get("results", [])
        filtered_pages = [p for p in pages if filter_page(p, modified_after, modified_before, created_in_years, modified_in_years)]
        for page in filtered_pages:
            page_id = page.get("id")
            full_data = safe_request(
                f"{base_url}/rest/api/content/{page_id}",
                headers, auth, params={"expand": "body.storage"}
            )
            if full_data:
                page["body"] = full_data.get("body", {})
                all_pages.append(page)
        if "_links" in data and "next" in data["_links"]:
            start += limit
        else:
            break
    return all_pages

def filter_page(page, modified_after, modified_before, created_in_years, modified_in_years=None):
    """Filter pages based on modification and creation dates.
    
    Args:
        created_in_years: None, single year (int), or list of years ([int])
        modified_in_years: None, single year (int), or list of years ([int])
    """
    created_date_str = page.get("history", {}).get("createdDate")
    modified_date_str = page.get("version", {}).get("when")
    
    created_date = parse_iso_date(created_date_str) if created_date_str else None
    modified_date = parse_iso_date(modified_date_str) if modified_date_str else None
    
    if modified_after and modified_date and modified_date < modified_after:
        return False
    if modified_before and modified_date and modified_date > modified_before:
        return False
    if created_in_years and created_date:
        if isinstance(created_in_years, list):
            if created_date.year not in created_in_years:
                return False
        elif created_date.year != created_in_years:
            return False
    if modified_in_years and modified_date:
        if isinstance(modified_in_years, list):
            if modified_date.year not in modified_in_years:
                return False
        elif modified_date.year != modified_in_years:
            return False
    return True

def parse_iso_date(date_str):
    """Parse ISO date string to datetime."""
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00'))
    except:
        return None

def get_last_editor_email(base_url, auth, headers, page_id):
    """Retrieve email of the last editor of a page."""
    try:
        url = f"{base_url}/rest/api/content/{page_id}/history"
        data = safe_request(url, headers, auth)
        if not data:
            return "?"
        
        last_updated = data.get("lastUpdated", {})
        by_user = last_updated.get("by", {})
        email = by_user.get("email", "?")
        return email if email else "?"
    except Exception as e:
        logging.error(f"Error fetching editor email for page {page_id}: {e}")
        return "?"

def get_attachments(base_url, auth, headers, page_id, created_in_years=None, modified_in_years=None):
    """Retrieve attachments for a given page.
    
    Args:
        created_in_years: None, single year (int), or list of years ([int])
        modified_in_years: None, single year (int), or list of years ([int])
    """
    try:
        url = f"{base_url}/rest/api/content/{page_id}/child/attachment"
        data = safe_request(url, headers, auth, params={"expand": "version,history,extensions,metadata"})
        if not data:
            return []
        
        attachments = data.get("results", [])
        
        # If no filter, return all
        if not created_in_years and not modified_in_years:
            return attachments
        
        filtered = []
        for att in attachments:
            created_str = att.get("history", {}).get("createdDate") or att.get("version", {}).get("when", "")
            modified_str = att.get("version", {}).get("when", "")
            created_date  = parse_iso_date(created_str)
            modified_date = parse_iso_date(modified_str)
            
            if created_in_years and created_date:
                if isinstance(created_in_years, list):
                    if created_date.year not in created_in_years:
                        continue
                elif created_date.year != created_in_years:
                    continue
            
            if modified_in_years and modified_date:
                if isinstance(modified_in_years, list):
                    if modified_date.year not in modified_in_years:
                        continue
                elif modified_date.year != modified_in_years:
                    continue
            
            filtered.append(att)
        
        return filtered
        
    except Exception as e:
        logging.error(f"Error fetching attachments for page {page_id}: {e}")
        return []

def load_keywords(keywords_file):
    """Load keywords from a file."""
    if not keywords_file or not os.path.exists(keywords_file):
        return []
    with open(keywords_file, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip() and not line.startswith("#")]

def load_patterns(regex_file, single_regex):
    """Load regex patterns from file or single pattern."""
    patterns = []
    if regex_file and os.path.exists(regex_file):
        with open(regex_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and ":::" in line:
                    parts = line.split(":::")
                    if len(parts) >= 3:
                        name = parts[0].strip()
                        regex = parts[1].strip()
                        group_index = int(parts[2].strip())
                        patterns.append((name, regex, group_index))
    elif single_regex:
        patterns.append(("custom_regex", single_regex, 0))
    return patterns

def load_trufflehog_patterns(yaml_file, filter_keywords=None, exclude_keywords=None):
    """
    Load regex patterns from TruffleHog YAML format.
    
    Format:
    - name: AWS API Key
      keywords:
      - AWS
      - API
      regex:
        AWS API Key: AKIA[0-9A-Z]{16}
    
    Args:
        yaml_file: Path to TruffleHog YAML file
        filter_keywords: Optional list of keywords to include detectors by.
                         A detector is included if ANY of its YAML keywords
                         matches ANY of the filter keywords (case-insensitive).
                         If None — all detectors are loaded (unless excluded).
        exclude_keywords: Optional list of keywords to exclude detectors by.
                          A detector is excluded if ANY of its YAML keywords
                          matches ANY of the exclude keywords (case-insensitive).
                          Applied after filter_keywords. Cannot be combined with
                          filter_keywords (mutually exclusive).
    
    Returns:
        List of tuples: (name, regex, group_index)
    """
    if not yaml:
        logging.error("PyYAML not installed. Install it with: pip install pyyaml --break-system-packages")
        return []
    
    if not os.path.exists(yaml_file):
        logging.error(f"TruffleHog YAML file not found: {yaml_file}")
        return []
    
    # Normalize to lowercase sets for case-insensitive comparison
    filter_kw_lower  = set(k.lower() for k in filter_keywords)  if filter_keywords  else None
    exclude_kw_lower = set(k.lower() for k in exclude_keywords) if exclude_keywords else None
    
    patterns = []
    skipped_include = 0
    skipped_exclude = 0
    
    try:
        with open(yaml_file, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
        
        if not data:
            logging.error(f"Empty or invalid YAML file: {yaml_file}")
            return []
        
        total = len([d for d in data if isinstance(d, dict)])
        
        for detector in data:
            if not isinstance(detector, dict):
                continue
            
            detector_name    = detector.get('name', 'Unknown')
            detector_keywords = [kw.lower() for kw in detector.get('keywords', [])]
            regex_dict       = detector.get('regex', {})
            
            # --- include filter: skip if no keyword matches ---
            if filter_kw_lower is not None:
                if not any(kw in filter_kw_lower for kw in detector_keywords):
                    skipped_include += 1
                    continue
            
            # --- exclude filter: skip if any keyword matches ---
            if exclude_kw_lower is not None:
                if any(kw in exclude_kw_lower for kw in detector_keywords):
                    skipped_exclude += 1
                    continue
            
            for pattern_name, pattern_regex in regex_dict.items():
                full_name = pattern_name if pattern_name == detector_name else f"{detector_name} - {pattern_name}"
                patterns.append((full_name, pattern_regex, 0))
        
        # --- summary log ---
        if filter_kw_lower is not None:
            logging.info(f"TruffleHog include filter : {sorted(filter_kw_lower)}")
            logging.info(f"  Detectors matched : {total - skipped_include}/{total}, skipped: {skipped_include}")
        
        if exclude_kw_lower is not None:
            logging.info(f"TruffleHog exclude filter : {sorted(exclude_kw_lower)}")
            logging.info(f"  Detectors excluded: {skipped_exclude}/{total}")
        
        logging.info(f"Loaded {len(patterns)} patterns from TruffleHog YAML: {yaml_file}")
        
        if patterns:
            logging.info("Sample patterns loaded:")
            for i, (name, regex, _) in enumerate(patterns[:5], 1):
                logging.info(f"  {i}. {name}: {regex[:60]}...")
        else:
            hint = sorted(filter_kw_lower or exclude_kw_lower or [])
            logging.warning(
                f"No patterns loaded after applying filters {hint}. "
                f"Check that keywords match those in your YAML file."
            )
        
        return patterns
    
    except yaml.YAMLError as e:
        logging.error(f"Error parsing TruffleHog YAML file: {e}")
        return []
    except Exception as e:
        logging.error(f"Error loading TruffleHog YAML file: {e}")
        return []


def scan_text_for_keywords(text, keywords):
    """Scan text for keywords (case-insensitive)."""
    findings = []
    for keyword in keywords:
        if re.search(re.escape(keyword), text, re.IGNORECASE):
            findings.append(("keyword", keyword, keyword))
    return findings

def scan_text_for_secrets(text, patterns):
    """Scan text for secrets using regex patterns."""
    findings = []
    for pattern_name, pattern_regex, group_index in patterns:
        try:
            matches = re.finditer(pattern_regex, text, re.MULTILINE | re.IGNORECASE)
            for match in matches:
                if group_index < len(match.groups()) + 1:
                    secret_value = match.group(group_index) if group_index > 0 else match.group(0)
                    findings.append((pattern_name, secret_value))
        except re.error as e:
            logging.error(f"Regex error in pattern '{pattern_name}': {e}")
            continue
    return findings

def sanitize_cell(value):
    """Remove NUL bytes and other characters that break CSV/XLSX writers."""
    return str(value).replace('\x00', '').replace('\r', ' ')

def format_secret_value(value, max_length=None):
    """Format secret value for display (truncate if needed)."""
    value_str = sanitize_cell(value).strip()
    if max_length is not None and len(value_str) > max_length:
        return value_str[:max_length] + "..."
    return value_str

def _try_download_attachment(url, auth, download_headers):
    """
    Try to download an attachment from a single URL, following redirects.
    Returns (response, error_code) where error_code is None on success.
    Uses download_attachment_with_redirect to handle Atlassian's 302→Media API flow.
    """
    return download_attachment_with_redirect(url, auth, download_headers)


def extract_text_from_attachment(base_url, auth, headers, attachment, max_size_bytes, keywords, patterns, scan_images_only, archive_support, cloud_id=None):
    """Extract text from various attachment types.

    Atlassian Cloud download flow (as of 2025):
      REST endpoint → HTTP 302 → api.media.atlassian.com/file/{id}/binary?token=...
    The token is embedded in the redirect URL so Basic Auth is NOT forwarded to the
    media host. download_attachment_with_redirect() handles this automatically.
    """
    att_title = attachment.get("title", "?")
    att_id = attachment.get("id", "")  # e.g. "att2649589343"

    file_size = attachment.get("extensions", {}).get("fileSize", 0)
    if max_size_bytes and file_size > max_size_bytes:
        logging.info(f"Skipping {att_title} (size: {file_size} bytes > max: {max_size_bytes})")
        return [], ""

    ext = os.path.splitext(att_title)[1][1:].lower()

    if scan_images_only and ext not in ["png", "jpg", "jpeg", "gif", "bmp", "tiff"]:
        return [], ext

    # Use Accept: */* for binary downloads
    download_headers = {k: v for k, v in headers.items() if k.lower() != "accept"}
    download_headers["Accept"] = "*/*"

    # Build candidate URL list — first working one wins.
    # Each URL will follow 302 redirects automatically (including to Media API).
    candidate_urls = []

    if att_id:
        # Best option: /rest/api/content/{page_id}/child/attachment/{att_id}/download
        # Confluence returns 302 → api.media.atlassian.com with a pre-signed token.
        container = attachment.get("_expandable", {}).get("container", "")
        page_id_match = re.search(r'/content/(\d+)', container)
        if page_id_match:
            page_id_from_container = page_id_match.group(1)
            candidate_urls.append(
                f"{base_url}/rest/api/content/{page_id_from_container}/child/attachment/{att_id}/download"
            )
        # self link + /download (same mechanism)
        self_link = attachment.get("_links", {}).get("self", "")
        if self_link:
            candidate_urls.append(f"{self_link}/download")

    # Legacy URL as final fallback (may 401 on some tenants but worth trying)
    legacy_url = f"{base_url}{attachment['_links'].get('download', '')}"
    if legacy_url not in candidate_urls:
        candidate_urls.append(legacy_url)

    try:
        response = None
        retry_delay = 10

        for url in candidate_urls:
            logging.debug(f"Trying download URL for {att_title}: {url}")
            resp, err_code = _try_download_attachment(url, auth, download_headers)

            if resp is not None:
                response = resp
                logging.debug(f"Downloaded {att_title} via: {url}")
                break
            elif err_code == 500:
                for attempt in range(2, 4):
                    logging.warning(f"500 for {att_title} (attempt {attempt}/3), retrying in {retry_delay}s...")
                    time.sleep(retry_delay)
                    resp, _ = _try_download_attachment(url, auth, download_headers)
                    if resp is not None:
                        response = resp
                        break
                if response:
                    break
                logging.error(f"500 persists for {att_title} at {url}, trying next URL...")
            else:
                logging.debug(f"HTTP {err_code} at {url} for {att_title}, trying next URL...")

        if response is None:
            logging.error(f"All download URLs failed for {att_title}. Tried: {candidate_urls}")
            return [], ext
        
        content = response.content
        
        text = ""
        
        # Text files (expanded list including scripts and code files)
        text_extensions = [
            # Basic text
            "txt", "log", "md", "markdown", "rst",
            # Config files
            "conf", "cfg", "ini", "properties", "env",
            # Data formats
            "json", "xml", "yaml", "yml", "toml", "csv", "tsv",
            # Scripts and code
            "py", "sh", "bash", "zsh", "fish", "bat", "cmd", "ps1",
            "js", "ts", "jsx", "tsx", "java", "c", "cpp", "h", "hpp",
            "cs", "go", "rs", "rb", "php", "pl", "r", "scala", "kt",
            "swift", "m", "mm", "dart", "lua", "groovy", "gradle",
            # Web
            "html", "htm", "css", "scss", "sass", "less",
            # Other
            "sql", "graphql", "proto", "dockerfile", "makefile",
            "terraform", "tf", "hcl", "jenkinsfile"
        ]
        
        if ext in text_extensions:
            text = content.decode("utf-8", errors="ignore")
        
        # DOCX
        elif ext == "docx" and docx:
            doc = docx.Document(BytesIO(content))
            text = "\n".join([para.text for para in doc.paragraphs])
        
        # PDF
        elif ext == "pdf" and fitz:
            pdf_doc = fitz.open(stream=content, filetype="pdf")
            text = "\n".join([page.get_text() for page in pdf_doc])
        
        # Images with OCR
        elif ext in ["png", "jpg", "jpeg", "gif", "bmp", "tiff"]:
            if not content:
                logging.warning(f"Skipping {att_title}: empty response body")
            elif content[:4] in (b'\x89PNG', b'\xff\xd8\xff') or content[:3] == b'GIF' or content[:2] == b'BM' or len(content) > 100:
                try:
                    img = Image.open(BytesIO(content))
                    img.verify()  # check integrity without decoding
                    img = Image.open(BytesIO(content))  # reopen after verify
                    text = pytesseract.image_to_string(img)
                except Exception as img_err:
                    logging.warning(f"Skipping {att_title}: not a valid image file ({img_err})")
            else:
                logging.warning(f"Skipping {att_title}: response doesn't look like an image ({len(content)} bytes)")
        
        # Archives (if enabled)
        elif archive_support and ext in ["zip", "tar", "gz", "tgz", "tar.gz"]:
            # Simple implementation: extract and scan text files
            if ext == "zip" and zipfile:
                with zipfile.ZipFile(BytesIO(content)) as zf:
                    for name in zf.namelist():
                        if name.endswith(('.txt', '.log', '.md', '.json', '.xml')):
                            text += zf.read(name).decode("utf-8", errors="ignore") + "\n"
        
        if not text:
            return [], ext
        
        # Strip NUL bytes and other problematic characters from extracted text
        # before scanning — prevents them from ending up in CSV/XLSX
        text = text.replace('\x00', '').replace('\r', ' ')
        
        # Scan for keywords and patterns
        findings = []
        if keywords:
            findings.extend(scan_text_for_keywords(text, keywords))
        if patterns:
            findings.extend([("regex", name, matched_text) for name, matched_text in scan_text_for_secrets(text, patterns)])
        
        return findings, ext
    
    except Exception as e:
        logging.error(f"Error extracting text from {att_title}: {e}")
        return [], ext

def parse_size(size_str):
    """Parse size string like '2mb' or '500kb' to bytes."""
    size_str = size_str.lower().strip()
    match = re.match(r"(\d+)(kb|mb|gb)?", size_str)
    if not match:
        return None
    num = int(match.group(1))
    unit = match.group(2) or "b"
    multipliers = {"b": 1, "kb": 1024, "mb": 1024**2, "gb": 1024**3}
    return num * multipliers.get(unit, 1)

def parse_date(date_str):
    """Parse date string in D.M.Y or D/M/Y format, returns timezone-aware datetime."""
    for sep in [".", "/"]:
        if sep in date_str:
            parts = date_str.split(sep)
            if len(parts) == 3:
                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                from datetime import timezone
                return datetime(year, month, day, tzinfo=timezone.utc)
    raise ValueError(f"Invalid date format: {date_str}")

def parse_age(age_str):
    """Parse age string like '1d', '1w', '1m', '1y' to timedelta."""
    age_str = age_str.lower().strip()
    match = re.match(r"(\d+)([dwmy])", age_str)
    if not match:
        return None
    num = int(match.group(1))
    unit = match.group(2)
    if unit == "d":
        return timedelta(days=num)
    elif unit == "w":
        return timedelta(weeks=num)
    elif unit == "m":
        return timedelta(days=num * 30)
    elif unit == "y":
        return timedelta(days=num * 365)
    return None

def load_list_from_arg(arg):
    """Load list from comma-separated string or file."""
    if not arg:
        return set()
    if os.path.isfile(arg):
        with open(arg, "r", encoding="utf-8") as f:
            return set([line.strip() for line in f if line.strip()])
    return set([k.strip() for k in arg.split(",")])

def validate_arguments(args):
    """Validate command-line arguments."""
    errors = []
    
    if not args.base_url:
        errors.append("--base-url is required")
    
    if not args.public_only:
        if not args.username:
            errors.append("--username is required (or use --public-only)")
        if not args.token:
            errors.append("--token is required (or use --public-only)")
    
    if not args.keywords and not args.regex_file and not args.regex and not getattr(args, 'trufflehog_yaml', None):
        errors.append("At least one of --keywords, --regex, --regex-file, or --trufflehog-patterns is required")
    
    # Validate mode parameter
    mode = getattr(args, 'mode', None)
    if mode in ['files', 'both']:
        if args.filetype and args.exclude_filetype:
            errors.append("Cannot use both --filetype and --exclude-filetype")
    
    if args.email_sender and not args.email_recipient:
        errors.append("--email-recipient is required when using --email-sender")
    
    if args.email_recipient and not args.email_sender:
        errors.append("--email-sender is required when using --email-recipient")
    
    # Validate alert flag
    if args.alert and not args.email_sender:
        errors.append("--email-sender is required when using --alert")
    
    # Validate trufflehog keyword filters require --trufflehog-patterns
    has_th_yaml    = bool(getattr(args, 'trufflehog_yaml', None))
    has_th_include = bool(getattr(args, 'trufflehog_keywords', None))
    has_th_exclude = bool(getattr(args, 'trufflehog_exclude_keywords', None))
    
    if has_th_include and not has_th_yaml:
        errors.append("--trufflehog-keywords requires --trufflehog-patterns to be specified")
    if has_th_exclude and not has_th_yaml:
        errors.append("--trufflehog-exclude-keywords requires --trufflehog-patterns to be specified")
    if has_th_include and has_th_exclude:
        errors.append("--trufflehog-keywords and --trufflehog-exclude-keywords are mutually exclusive — use one or the other")
    
    # --trufflehog-patterns is mutually exclusive with --regex-file, --regex and --keywords
    if has_th_yaml and args.regex_file:
        errors.append("--trufflehog-patterns and --regex-file are mutually exclusive — use one or the other")
    if has_th_yaml and args.regex:
        errors.append("--trufflehog-patterns and --regex are mutually exclusive — use one or the other")
    if has_th_yaml and args.keywords:
        errors.append("--trufflehog-patterns and --keywords are mutually exclusive — use one or the other")
    
    return errors

def load_config(config_file):
    """Load configuration from JSON file."""
    if not config_file or not os.path.exists(config_file):
        return {}
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Error loading config file: {e}")
        return {}

def create_xlsx_report(csv_file, xlsx_file):
    """Convert CSV to formatted XLSX report with improved formatting."""
    if not Workbook:
        logging.warning("openpyxl not installed, skipping XLSX creation")
        return False
    
    try:
        # Read CSV and strip NUL bytes that would break openpyxl
        with open(csv_file, 'r', encoding='utf-8', errors='replace') as f:
            content_clean = f.read().replace('\x00', '')
        reader = csv.reader(content_clean.splitlines())
        rows = list(reader)
        
        if len(rows) < 2:
            logging.warning("No data to create XLSX report")
            return False
        
        # Remove dummy row if present
        if len(rows) > 1 and "DUMMY_ROW_DELETE_ME" in rows[1][0]:
            rows.pop(1)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Confluence Secrets"
        
        # Define styles
        header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write data
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                
                # Header styling (first row)
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Secret value in red (last column or "Matched Value")
                    if col_idx == len(row_data) or (row_idx == 1 and 'Matched Value' in str(value)):
                        if row_idx > 1:  # Don't color header
                            cell.font = Font(color='DC143C', bold=True)
                    
                    # Make URLs clickable
                    if 'http' in str(value).lower() and '://' in str(value):
                        cell.hyperlink = value
                        cell.font = Font(color='0563C1', underline='single')
        
        # Auto-adjust column widths based on content
        for col_idx in range(1, len(rows[0]) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Check header
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value:
                max_length = len(str(header_value))
            
            # Check first 100 rows for performance
            for row_idx in range(2, min(len(rows) + 1, 102)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # For URLs, limit the length calculation
                    if 'http' in str(cell_value).lower():
                        max_length = max(max_length, min(len(str(cell_value)), 60))
                    else:
                        max_length = max(max_length, min(len(str(cell_value)), 50))
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 70)
            if adjusted_width < 15:
                adjusted_width = 15
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set header row height
        ws.row_dimensions[1].height = 30
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save workbook
        wb.save(xlsx_file)
        logging.info(f"XLSX report created: {xlsx_file}")
        return True
        
    except Exception as e:
        logging.error(f"Error creating XLSX report: {e}")
        return False

def create_author_report(author_findings, author_email, filename):
    """
    Create personalized Excel report for a specific author
    
    Args:
        author_findings: List of findings for this author
        author_email: Author's email address
        filename: Output filename for the report
    
    Returns:
        Filename of created report
    """
    if not Workbook:
        logging.error("openpyxl not installed, cannot create author report")
        return None
    
    try:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Your Confluence Secrets"
        
        # Define border
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Check if any findings have attachment info
        has_attachments = any(f.get('is_attachment', False) for f in author_findings)
        
        # Row 1: Main title with red background
        if has_attachments:
            headers = ['Space Name', 'Page Title', 'File Name', 'File Type', 'File URL', 'Page URL', 'Matched Keyword', 'Finding Type', 'Matched Value']
        else:
            headers = ['Space Name', 'Page Title', 'Page URL', 'Matched Keyword', 'Finding Type', 'Matched Value']
        
        last_col = get_column_letter(len(headers))
        
        sheet.merge_cells(f'A1:{last_col}1')
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = "SECURITY ALERT: Exposed Secrets Found in Your Confluence Pages"
        title_cell.font = Font(size=14, bold=True, color='FFFFFF')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        sheet.row_dimensions[1].height = 25
        
        # Row 2: Warning message with yellow background
        sheet.merge_cells(f'A2:{last_col}2')
        warning_cell = sheet.cell(row=2, column=1)
        warning_cell.value = "The following secrets were detected in pages you edited. Please take immediate action."
        warning_cell.font = Font(size=10, italic=True)
        warning_cell.alignment = Alignment(horizontal='center', vertical='center')
        warning_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        sheet.row_dimensions[2].height = 20
        
        # Row 3: Empty row for spacing
        sheet.row_dimensions[3].height = 5
        
        # Row 4: Column headers with red background
        header_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        sheet.row_dimensions[4].height = 30
        
        # Add findings (starting from row 5)
        for row_num, finding in enumerate(author_findings, 5):
            col = 1
            
            # Space Name
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('space_name', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Page Title
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('page_title', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            if has_attachments:
                # File Name (only if this finding has attachment info)
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    cell.value = finding.get('file_title', 'N/A')
                else:
                    cell.value = '(Page Content)'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
                
                # File Type
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    cell.value = finding.get('file_extension', 'N/A')
                else:
                    cell.value = 'N/A'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
                
                # File URL
                cell = sheet.cell(row=row_num, column=col)
                if finding.get('is_attachment', False):
                    file_url = finding.get('file_url', '')
                    cell.value = file_url
                    if file_url:
                        cell.hyperlink = file_url
                        cell.font = Font(color='0563C1', underline='single')
                else:
                    cell.value = 'N/A'
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                col += 1
            
            # Page URL as hyperlink
            url_cell = sheet.cell(row=row_num, column=col)
            page_url = finding.get('page_url', '')
            url_cell.value = page_url
            if page_url:
                url_cell.hyperlink = page_url
                url_cell.font = Font(color='0563C1', underline='single')
            url_cell.border = border
            url_cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Matched Keyword
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('keyword', 'Unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Finding Type
            cell = sheet.cell(row=row_num, column=col)
            cell.value = finding.get('finding_type', 'unknown')
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            col += 1
            
            # Secret value - red bold font
            secret_cell = sheet.cell(row=row_num, column=col)
            secret_cell.value = finding.get('matched_value', '')
            secret_cell.font = Font(color='DC143C', bold=True, size=10)
            secret_cell.border = border
            secret_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths based on content
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length for this column
            max_length = len(header)
            for row_num in range(5, 5 + len(author_findings)):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    # For URLs and long text, limit the calculation
                    if 'URL' in header or 'http' in str(cell_value).lower():
                        max_length = max(max_length, min(len(str(cell_value)), 60))
                    else:
                        max_length = max(max_length, min(len(str(cell_value)), 50))
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 70)
            if adjusted_width < 15:
                adjusted_width = 15
            
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top rows (header and title)
        sheet.freeze_panes = "A5"
        
        wb.save(filename)
        return filename
    
    except Exception as e:
        logging.error(f"Error creating author report for {author_email}: {e}")
        return None


def send_author_alert(author_email, author_name, author_findings, aws_region='eu-central-1', 
                     sender_email=None, security_contact='security@company.com', 
                     security_wiki=None, has_attachments=False):
    """
    Send personalized alert email to page editor
    
    Args:
        author_email: Editor's email address
        author_name: Editor's display name (can be email if name not available)
        author_findings: List of secrets found in pages edited by this person
        aws_region: AWS region for SES
        sender_email: Sender email address
        security_contact: Security team contact email
        security_wiki: Security documentation URL (optional)
        has_attachments: Boolean indicating if scan included attachments
    
    Returns:
        Boolean indicating success
    """
    if not boto3:
        logging.error(f"      ❌ boto3 not available for {author_email}")
        return False
    
    if not sender_email:
        logging.error(f"      ❌ Sender email not specified for {author_email}")
        return False
    
    try:
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Create author-specific report
        temp_filename = f"temp_author_{author_email.replace('@', '_at_').replace('.', '_')}.xlsx"
        create_author_report(author_findings, author_email, temp_filename)
        
        # Prepare subject
        secret_count = len(author_findings)
        page_count = len(set(f['page_url'] for f in author_findings))
        subject = f"SECURITY ALERT: {secret_count} Secret{'s' if secret_count != 1 else ''} Found in Your Confluence Page{'s' if page_count != 1 else ''}"
        
        # Prepare email body
        body_text = f"""Hello {author_name if author_name and author_name != author_email else 'there'},

We have detected {secret_count} exposed secret{'s' if secret_count != 1 else ''} in {page_count} Confluence page{'s' if page_count != 1 else ''} that you last edited.

AFFECTED PAGES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        
        # List affected pages
        for page_url in sorted(set(f['page_url'] for f in author_findings)):
            page_findings = [f for f in author_findings if f['page_url'] == page_url]
            page_title = page_findings[0].get('page_title', 'Unknown')
            body_text += f"\n• {page_title}: {len(page_findings)} secret{'s' if len(page_findings) != 1 else ''} found\n"
            body_text += f"  {page_url}\n"
        
        body_text += f"""
IMMEDIATE ACTION REQUIRED
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. Review the attached Excel report for complete details
2. Rotate/revoke ALL exposed credentials immediately
3. Remove secrets from the affected Confluence pages{' and attachments' if has_attachments else ''}
4. Update applications with new credentials
5. Never store secrets in Confluence pages, comments, or attachments

The detailed report is attached to this email. Please address these issues immediately.

If you have any questions, contact your security team: {security_contact}
"""
        
        if security_wiki:
            body_text += f"\nSecurity documentation: {security_wiki}\n"
        
        body_text += """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
This is an automated security alert from Confluence Secrets Scanner.
"""
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = author_email
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain"))
        
        # Determine attachment filename based on scan mode
        if has_attachments:
            attachment_filename = "your_confluence_secrets_in_files.xlsx"
        else:
            attachment_filename = "your_confluence_secrets.xlsx"
        
        # Attach Excel report with appropriate name
        if os.path.exists(temp_filename):
            with open(temp_filename, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename= {attachment_filename}",
            )
            msg.attach(part)
        
        # Send email
        response = ses_client.send_raw_email(
            Source=sender_email,
            Destinations=[author_email],
            RawMessage={"Data": msg.as_string()}
        )
        
        # Clean up temp file
        if os.path.exists(temp_filename):
            os.remove(temp_filename)
        
        logging.info(f"      ✅ Alert sent to {author_name} ({author_email})")
        return True
        
    except Exception as e:
        logging.error(f"      ❌ Failed to send alert to {author_email}: {e}")
        # Clean up temp file on error
        if os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except:
                pass
        return False


def generate_email_summary(total_secrets, total_pages, total_files, spaces_scanned, duration_str, include_attachments, affected_spaces=None, affected_pages=None):
    """Generate email summary in the requested format"""
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    
    # Use actual affected counts if provided, otherwise use scanned counts
    actual_spaces = affected_spaces if affected_spaces is not None else spaces_scanned
    actual_pages = affected_pages if affected_pages is not None else total_pages
    
    # Build email body
    body = f"""Confluence Secrets Scanner Report - {timestamp}

Summary Statistics:
* Total Secrets Found: {total_secrets}
* Affected Spaces: {actual_spaces}
* Affected Pages: {actual_pages}"""
    
    if include_attachments:
        body += f"\n* Files Scanned: {total_files}"
    
    body += f"""

ACTION REQUIRED:
1. Review the attached report immediately
2. Rotate/revoke exposed credentials
3. Implement proper secrets management

The detailed report is attached as XLSX file.

---
This is an automated report generated by Confluence Secrets Scanner.
"""
    
    return body

def count_affected_from_csv(csv_file):
    """Count unique affected spaces and pages from CSV results"""
    try:
        if not os.path.exists(csv_file):
            return None, None
        
        unique_spaces = set()
        unique_pages = set()
        
        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            
            for row in reader:
                if len(row) >= 3:  # Need at least space_name, title, page_url
                    space_name = row[0].strip()
                    page_url = row[2].strip() if len(row) > 2 else ""
                    
                    if space_name:
                        unique_spaces.add(space_name)
                    if page_url:
                        unique_pages.add(page_url)
        
        return len(unique_spaces), len(unique_pages)
    
    except Exception as e:
        logging.warning(f"Could not count affected spaces/pages from CSV: {e}")
        return None, None

def send_email_with_attachment(subject, body_text, sender, recipient, aws_region, attachment_path):
    """Send email via AWS SES with XLSX attachment
    
    Args:
        recipient: String with single email or comma-separated emails (e.g., "user1@example.com,user2@example.com")
    """
    if not boto3:
        logging.error("boto3 not installed. Install with: pip install boto3")
        return False
    
    try:
        # Parse recipients - support both single email and comma-separated list
        if isinstance(recipient, str):
            recipients = [email.strip() for email in recipient.split(',') if email.strip()]
        else:
            recipients = [recipient]
        
        if not recipients:
            logging.error("No valid recipients specified")
            return False
        
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)  # Join all recipients for the header
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain", "utf-8"))
        
        # Attach XLSX file
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as f:
                attachment = MIMEApplication(f.read(), _subtype="xlsx")
                attachment.add_header(
                    "Content-Disposition", 
                    "attachment", 
                    filename=os.path.basename(attachment_path)
                )
                msg.attach(attachment)
            logging.info(f"Attached file: {attachment_path}")
        
        # Send email to all recipients
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=recipients,  # AWS SES expects a list
            RawMessage={"Data": msg.as_string()}
        )
        
        logging.info(f"✅ Email sent successfully to {len(recipients)} recipient(s): {', '.join(recipients)}")
        logging.info(f"   MessageId: {response['MessageId']}")
        return True
        
    except Exception as e:
        logging.error(f"❌ Error sending email: {e}")
        return False

def send_email_with_multiple_attachments(subject, body_text, sender, recipient, aws_region, attachment_paths):
    """Send email via AWS SES with multiple XLSX attachments
    
    Args:
        recipient: String with single email or comma-separated emails
        attachment_paths: List of file paths to attach
    """
    if not boto3:
        logging.error("boto3 not installed. Install with: pip install boto3")
        return False
    
    try:
        # Parse recipients - support both single email and comma-separated list
        if isinstance(recipient, str):
            recipients = [email.strip() for email in recipient.split(',') if email.strip()]
        else:
            recipients = [recipient]
        
        if not recipients:
            logging.error("No valid recipients specified")
            return False
        
        # Create SES client
        ses_client = boto3.client("ses", region_name=aws_region)
        
        # Create email message
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        
        # Attach body
        msg.attach(MIMEText(body_text, "plain", "utf-8"))
        
        # Attach multiple XLSX files
        for attachment_path in attachment_paths:
            if os.path.exists(attachment_path):
                with open(attachment_path, "rb") as f:
                    attachment = MIMEApplication(f.read(), _subtype="xlsx")
                    attachment.add_header(
                        "Content-Disposition", 
                        "attachment", 
                        filename=os.path.basename(attachment_path)
                    )
                    msg.attach(attachment)
                logging.info(f"Attached file: {attachment_path}")
            else:
                logging.warning(f"Attachment not found: {attachment_path}")
        
        # Send email to all recipients
        response = ses_client.send_raw_email(
            Source=sender,
            Destinations=recipients,
            RawMessage={"Data": msg.as_string()}
        )
        
        logging.info(f"✅ Email sent successfully to {len(recipients)} recipient(s): {', '.join(recipients)}")
        logging.info(f"   MessageId: {response['MessageId']}")
        logging.info(f"   Attachments: {len(attachment_paths)} file(s)")
        return True
        
    except Exception as e:
        logging.error(f"❌ Error sending email: {e}")
        return False

def validate_csv_fields(space_name, title, page_url, keyword, finding_type, formatted_value, email):
    """
    Validate and sanitize all fields before writing to CSV.
    Ensures all fields are non-empty strings to prevent column shifting.
    """
    safe_space_name = sanitize_cell(space_name).strip() if space_name else "Unknown Space"
    safe_title = sanitize_cell(title).strip() if title and str(title).strip() != "?" else "Untitled Page"
    safe_page_url = sanitize_cell(page_url).strip() if page_url else "No URL"
    safe_keyword = sanitize_cell(keyword).strip() if keyword else "Unknown Keyword"
    safe_finding_type = sanitize_cell(finding_type).strip() if finding_type else "unknown"
    safe_formatted_value = sanitize_cell(formatted_value).strip() if formatted_value else "No Value"
    safe_email = sanitize_cell(email).strip() if email and str(email).strip() != "?" else "unknown@unknown.com"
    
    # Log warning if any field was replaced with default
    if not space_name or not str(space_name).strip():
        logging.warning(f"Empty space_name, using default: {safe_space_name}")
    if not title or str(title).strip() == "?":
        logging.warning(f"Empty or missing title for URL {safe_page_url}, using default: {safe_title}")
    if not email or str(email).strip() == "?":
        logging.warning(f"Empty or missing email for {safe_title}, using default: {safe_email}")
    
    return (safe_space_name, safe_title, safe_page_url, safe_keyword, 
            safe_finding_type, safe_formatted_value, safe_email)

def validate_attachment_fields(space_name, page_title, file_title, ext, file_url, page_url, keyword, finding_type, formatted_value, email):
    """
    Validate and sanitize attachment fields before writing to CSV.
    Ensures all 10 fields are non-empty strings to prevent column shifting.
    """
    safe_space_name = sanitize_cell(space_name).strip() if space_name else "Unknown Space"
    safe_page_title = sanitize_cell(page_title).strip() if page_title and str(page_title).strip() != "?" else "Untitled Page"
    safe_file_title = sanitize_cell(file_title).strip() if file_title and str(file_title).strip() != "?" else "Unknown File"
    safe_ext = sanitize_cell(ext).strip() if ext else "unknown"
    safe_file_url = sanitize_cell(file_url).strip() if file_url else "No URL"
    safe_page_url = sanitize_cell(page_url).strip() if page_url else "No URL"
    safe_keyword = sanitize_cell(keyword).strip() if keyword else "Unknown Keyword"
    safe_finding_type = sanitize_cell(finding_type).strip() if finding_type else "unknown"
    safe_formatted_value = sanitize_cell(formatted_value).strip() if formatted_value else "No Value"
    safe_email = sanitize_cell(email).strip() if email and str(email).strip() != "?" else "unknown@unknown.com"
    
    # Log warnings
    if not page_title or str(page_title).strip() == "?":
        logging.warning(f"Empty page title for attachment {safe_file_title}")
    if not file_title or str(file_title).strip() == "?":
        logging.warning(f"Empty file title in page {safe_page_title}")
    
    return (safe_space_name, safe_page_title, safe_file_title, safe_ext, safe_file_url,
            safe_page_url, safe_keyword, safe_finding_type, safe_formatted_value, safe_email)

def process_space(base_url, auth, headers, space, keywords, patterns, writer, csvfile, include_attachments, allowed_types, excluded_types, max_size_bytes, mod_after, mod_before, created_years, modified_in_years, no_duplicates, secret_max_length, scan_images_only, archive_support, findings_set, debug_limit=None, current_total=0, cloud_id=None):
    """Process a single Confluence space."""
    space_key = space.get("key", "?")
    space_name = space.get("name", space_key).strip()
    logging.info(f"Processing space: {space_key} - {space_name}")
    pages = get_pages_in_space(base_url, auth, headers, space_key, mod_after, mod_before, created_years, modified_in_years)
    
    total_files_in_space = 0
    total_secrets_in_space = 0
    total_pages_in_space = len(pages)
    
    # Don't reopen file - use existing writer
    for page in pages:
        secrets_found, files_scanned = process_page(base_url, auth, headers, page, space_name, space_key, writer, csvfile, keywords, patterns, include_attachments, allowed_types, excluded_types, max_size_bytes, no_duplicates, secret_max_length,
            scan_images_only, archive_support, findings_set, created_years, modified_in_years, cloud_id=cloud_id
        )
        total_secrets_in_space += secrets_found
        total_files_in_space += files_scanned
        
        # Debug mode: stop if reached limit
        if debug_limit and (current_total + total_secrets_in_space) >= debug_limit:
            logging.warning(f"🔍 DEBUG MODE: Reached limit in space '{space_name}'. Stopping...")
            break
    
    summary = f"[SUMMARY] Space '{space_name}': {total_pages_in_space} pages scanned, {total_secrets_in_space} secrets found"
    if include_attachments:
        summary += f", {total_files_in_space} files scanned"
    logging.info(summary)
    
    return total_secrets_in_space, total_files_in_space, total_pages_in_space

def process_page(base_url, auth, headers, page, space_name, space_key, writer, csvfile, keywords, patterns, include_attachments, allowed_types, excluded_types, max_size_bytes, no_duplicates, secret_max_length, scan_images_only, archive_support, findings_set, created_years=None, modified_in_years=None, cloud_id=None):
    """Process a single Confluence page."""
    title = page.get("title", "?")
    page_id = page.get("id", "?")
    page_url = f"{base_url}/spaces/{space_key}/pages/{page_id}"
    total_files_in_page = 0
    secrets_found = 0
    
    if include_attachments:
        attachments = get_attachments(base_url, auth, headers, page_id, created_years, modified_in_years)
        total_files_in_page = len(attachments)
        filtered_attachments = [
            att for att in attachments
            if (allowed_types is None or os.path.splitext(att.get("title", ""))[1][1:].lower() in allowed_types)
            and os.path.splitext(att.get("title", ""))[1][1:].lower() not in excluded_types
        ]
        for att in filtered_attachments:
            file_findings, ext = extract_text_from_attachment(
                base_url, auth, headers, att, max_size_bytes, keywords, patterns,
                scan_images_only, archive_support, cloud_id=cloud_id
            )
            if not file_findings:
                continue
            for finding_type, keyword, matched_text in file_findings:
                formatted_value = format_secret_value(matched_text, max_length=secret_max_length)
                # Use REST API download URL (Basic Auth compatible).
                # Legacy /download/attachments/... now requires OAuth on Atlassian Cloud.
                att_id = att.get("id", "")
                if att_id:
                    file_url = f"{base_url}/rest/api/content/{att_id}/download"
                else:
                    file_url = f"{base_url}{att['_links'].get('download', '')}"
                email = get_last_editor_email(base_url, auth, headers, page_id)
                
                # Validate all attachment fields before writing to CSV
                finding_tuple = validate_attachment_fields(
                    space_name, title, att.get("title", "?"), ext, 
                    file_url, page_url, keyword, finding_type, 
                    formatted_value, email
                )
                
                if no_duplicates and finding_tuple in findings_set:
                    continue
                if no_duplicates:
                    findings_set.add(finding_tuple)
                logging.info(f"[FOUND] Page: {title}, File: {att.get('title', '?')}, Keyword: {keyword}, Type: {finding_type}, Value: {formatted_value}")
                writer.writerow(list(finding_tuple))
                csvfile.flush()
                secrets_found += 1
    else:
        body = page.get("body", {}).get("storage", {}).get("value", "")
        findings = []
        if keywords:
            findings.extend(scan_text_for_keywords(body, keywords))
        if patterns:
            findings.extend([("regex", name, matched_text) for name, matched_text in scan_text_for_secrets(body, patterns)])
        for finding_type, keyword, matched_text in findings:
            formatted_value = format_secret_value(matched_text, max_length=secret_max_length)
            email = get_last_editor_email(base_url, auth, headers, page_id)
            
            # Validate all fields before writing to CSV
            finding_tuple = validate_csv_fields(
                space_name, title, page_url, keyword, 
                finding_type, formatted_value, email
            )
            
            if no_duplicates and finding_tuple in findings_set:
                continue
            if no_duplicates:
                findings_set.add(finding_tuple)
            logging.info(f"[FOUND] Page: {title}, Keyword: {keyword}, Type: {finding_type}, Value: {formatted_value}")
            writer.writerow(list(finding_tuple))
            csvfile.flush()
            secrets_found += 1
    
    return secrets_found, total_files_in_page

def export_findings_to_json(csv_file, json_file, include_attachments, scan_stats=None):
    """
    Read the CSV output and export findings to a structured JSON file.

    Args:
        csv_file:            Path to the CSV produced by the scan.
        json_file:           Destination JSON path.
        include_attachments: Whether the scan included attachments (affects CSV columns).
        scan_stats:          Optional dict with summary statistics to embed in the JSON.
    """
    findings = []
    try:
        with open(csv_file, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Skip the dummy row used to work-around the first-row bug
                values = list(row.values())
                if values and "DUMMY_ROW_DELETE_ME" in str(values[0]):
                    continue
                findings.append(dict(row))
    except Exception as e:
        logging.error(f"Error reading CSV for JSON export: {e}")

    output = {
        "generated_at": datetime.now().isoformat(),
        "scan_stats": scan_stats or {},
        "total_findings": len(findings),
        "findings": findings,
    }

    try:
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False, default=str)
        logging.info(f"JSON report created: {json_file}")
    except Exception as e:
        logging.error(f"Error writing JSON report: {e}")


def main(base_url, username, token, keywords_file, regex_file, single_regex, output_file,
         filetypes, exclude_filetypes, max_size, public_only,
         space_keys, exclude_space_keys, modified_after, modified_before, created_in_year,
         modified_in_year, no_duplicates, resume_from, scan_images_only, archive_support,
         config, email_sender, email_recipient, aws_region, secret_max_length, debug_mode,
         alert, security_contact, security_wiki, mode=None, trufflehog_yaml=None,
         trufflehog_keywords=None, trufflehog_exclude_keywords=None, _recursive=False,
         export_json=False):
    """Main function to scan Confluence for keywords and secrets."""
    start_time = time.time()
    
    # Only setup logging on first (non-recursive) call
    if not _recursive:
        setup_logging()
        
        # Remove previous scan results before starting
        output_dir = os.path.dirname(os.path.abspath(output_file))
        base_output = output_file.replace('.csv', '')
        files_to_clean = [
            output_file,
            f"{base_output}_pages.csv",
            f"{base_output}_files.csv",
            os.path.join(output_dir, "confluence_secrets.xlsx"),
            os.path.join(output_dir, "confluence_secrets_in_files.xlsx"),
            f"{base_output}.json",
            f"{base_output}_files.json",
            f"{base_output}_pages.json",
        ]
        for f in files_to_clean:
            if os.path.exists(f):
                try:
                    os.remove(f)
                    logging.info(f"Removed previous result: {os.path.basename(f)}")
                except Exception as e:
                    logging.warning(f"Could not remove {os.path.basename(f)}: {e}")
    
    # Handle mode parameter
    # If mode is None, scan only pages (default behavior)
    # If mode is 'both', run scan twice: once for pages, once for files
    if mode == "both":
        logging.info("=" * 80)
        logging.info("RUNNING IN BOTH MODE - Will scan pages and files separately")
        logging.info("=" * 80)
        
        # Generate separate output files for each phase
        base_output = output_file.replace('.csv', '')
        pages_csv = f"{base_output}_pages.csv"
        files_csv = f"{base_output}_files.csv"
        
        # Scan pages first (without sending email)
        logging.info("\n" + "=" * 80)
        logging.info("PHASE 1: SCANNING PAGES")
        logging.info("=" * 80)
        main(base_url, username, token, keywords_file, regex_file, single_regex, pages_csv,
             filetypes, exclude_filetypes, max_size, public_only,
             space_keys, exclude_space_keys, modified_after, modified_before, created_in_year,
             modified_in_year, no_duplicates, resume_from, scan_images_only, archive_support,
             config, None, None, aws_region, secret_max_length, debug_mode,  # email_sender=None, email_recipient=None
             alert, security_contact, security_wiki, mode=None,
             trufflehog_yaml=trufflehog_yaml, trufflehog_keywords=trufflehog_keywords,
             trufflehog_exclude_keywords=trufflehog_exclude_keywords,
             _recursive=True, export_json=export_json)  # Recursive call
        
        # Scan files second (without sending email)
        logging.info("\n" + "=" * 80)
        logging.info("PHASE 2: SCANNING FILES")
        logging.info("=" * 80)
        main(base_url, username, token, keywords_file, regex_file, single_regex, files_csv,
             filetypes, exclude_filetypes, max_size, public_only,
             space_keys, exclude_space_keys, modified_after, modified_before, created_in_year,
             modified_in_year, no_duplicates, resume_from, scan_images_only, archive_support,
             config, None, None, aws_region, secret_max_length, debug_mode,  # email_sender=None, email_recipient=None
             alert, security_contact, security_wiki, mode="files",
             trufflehog_yaml=trufflehog_yaml, trufflehog_keywords=trufflehog_keywords,
             trufflehog_exclude_keywords=trufflehog_exclude_keywords,
             _recursive=True, export_json=export_json)  # Recursive call
        
        # Derive actual output filenames from -o base name
        output_base_both = re.sub(r'\.(csv|xlsx|json)$', '', output_file, flags=re.IGNORECASE)
        output_dir_both = os.path.dirname(os.path.abspath(output_file))

        # Internal names (generated by each recursive call)
        xlsx_pages_internal = output_base_both + "_pages.xlsx"
        xlsx_files_internal = output_base_both + "_files.xlsx"

        # Final names that will be sent by email and shown in logs
        xlsx_pages = os.path.join(output_dir_both, "confluence_secrets.xlsx")
        xlsx_files = os.path.join(output_dir_both, "confluence_secrets_in_files.xlsx")

        # Rename internal files to final names
        for src, dst in [(xlsx_pages_internal, xlsx_pages), (xlsx_files_internal, xlsx_files)]:
            if os.path.exists(src):
                try:
                    os.replace(src, dst)
                    logging.info(f"Renamed {os.path.basename(src)} → {os.path.basename(dst)}")
                except Exception as e:
                    logging.warning(f"Could not rename {src} to {dst}: {e}")
                    # Fall back to original name if rename failed
                    if src == xlsx_pages_internal:
                        xlsx_pages = src
                    else:
                        xlsx_files = src

        json_pages = output_base_both + "_pages.json"
        json_files = output_base_both + "_files.json"

        # Build completion log listing all created files
        created_files = []
        for f in [xlsx_pages, xlsx_files]:
            if os.path.exists(f):
                created_files.append(os.path.basename(f))
        if export_json:
            for f in [json_pages, json_files]:
                if os.path.exists(f):
                    created_files.append(os.path.basename(f))
        files_list = ", ".join(created_files) if created_files else "no output files found"
        logging.info("\n" + "=" * 80)
        logging.info(f"BOTH MODE COMPLETED - Reports: {files_list}")
        logging.info("=" * 80)

        # Now send ONE email with BOTH attachments if email is configured
        if email_sender and email_recipient:
            if not aws_region:
                aws_region = "eu-central-1"

            logging.info("Preparing to send combined email report with both attachments...")
            
            pages_exist = os.path.exists(xlsx_pages)
            files_exist = os.path.exists(xlsx_files)
            
            if not pages_exist and not files_exist:
                logging.warning("No XLSX reports found to attach")
                return
            
            # Count secrets from BOTH CSV files
            total_secrets_pages = 0
            total_secrets_files = 0
            
            # Count from pages CSV
            if os.path.exists(pages_csv):
                try:
                    with open(pages_csv, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        for row in reader:
                            if row and "DUMMY_ROW_DELETE_ME" not in str(row):
                                total_secrets_pages += 1
                except Exception as e:
                    logging.warning(f"Could not count secrets from pages CSV: {e}")
            
            # Count from files CSV
            if os.path.exists(files_csv):
                try:
                    with open(files_csv, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        for row in reader:
                            if row and "DUMMY_ROW_DELETE_ME" not in str(row):
                                total_secrets_files += 1
                except Exception as e:
                    logging.warning(f"Could not count secrets from files CSV: {e}")
            
            total_secrets_combined = total_secrets_pages + total_secrets_files
            
            # Generate subject
            if total_secrets_combined > 0:
                subject = f"CRITICAL: Confluence Secrets Scanner (Pages & Files) - {total_secrets_combined} Secrets Found"
            else:
                subject = "Confluence Secrets Scanner (Pages & Files) - No Secrets Found"
            
            # Generate email body
            email_body = f"""This report covered page content and file attachments.

Total Secrets Found: {total_secrets_combined}
- Pages: {total_secrets_pages}
- Files: {total_secrets_files}

Two separate reports are attached:
1. confluence_secrets.xlsx - Secrets found in page content
2. confluence_secrets_in_files.xlsx - Secrets found in file attachments

Please review both reports and take appropriate action.
"""
            
            # Prepare attachments list
            attachments = []
            if pages_exist:
                attachments.append(xlsx_pages)
            if files_exist:
                attachments.append(xlsx_files)
            
            # Send email with multiple attachments
            send_email_with_multiple_attachments(
                subject=subject,
                body_text=email_body,
                sender=email_sender,
                recipient=email_recipient,
                aws_region=aws_region,
                attachment_paths=attachments
            )
        else:
            logging.info("Email notification disabled (no sender/recipient configured)")
        
        return
    
    # Set include_attachments based on mode
    if mode == "files":
        include_attachments = True
        logging.info("Running in FILES mode - scanning only attachments")
    else:  # mode is None (default - pages only)
        include_attachments = False
        logging.info("Running in PAGES mode - scanning only page content")
    
    # Load config file and override unset arguments
    config_dict = load_config(config)
    
    # Check dependencies
    dep_errors = check_dependencies()
    if dep_errors and (mode == "files" or scan_images_only or archive_support):
        for error in dep_errors:
            logging.error(error)
        exit(1)
    
    base_url = normalize_base_url(base_url)
    logging.info(f"Using base URL: {base_url}")
    
    # Debug mode warning
    if debug_mode:
        logging.warning("=" * 80)
        logging.warning("🔍 DEBUG MODE ENABLED - Script will stop after first 5 findings!")
        logging.warning("=" * 80)
    
    if public_only:
        auth = None
        logging.info("Running in PUBLIC-ONLY mode (no authentication)")
    else:
        auth = HTTPBasicAuth(username, token)
        logging.info("Running in AUTHENTICATED mode")
    
    headers = {"Accept": "application/json"}
    keywords = load_keywords(keywords_file) if keywords_file else []
    patterns = load_patterns(regex_file, single_regex)
    
    # Load TruffleHog patterns if specified and merge with existing patterns
    if trufflehog_yaml:
        # Parse trufflehog_keywords from comma-separated string to list (if provided)
        th_filter  = None
        th_exclude = None
        if trufflehog_keywords:
            th_filter = [k.strip() for k in trufflehog_keywords.split(',') if k.strip()]
        if trufflehog_exclude_keywords:
            th_exclude = [k.strip() for k in trufflehog_exclude_keywords.split(',') if k.strip()]
        
        trufflehog_patterns = load_trufflehog_patterns(
            trufflehog_yaml,
            filter_keywords=th_filter,
            exclude_keywords=th_exclude
        )
        if trufflehog_patterns:
            patterns.extend(trufflehog_patterns)
            logging.info(f"Total patterns after merging TruffleHog: {len(patterns)}")
    
    allowed_types = set([t.lower() for t in filetypes.split(",")]) if filetypes else None
    excluded_types = set([t.lower() for t in exclude_filetypes.split(",")]) if exclude_filetypes else set()
    max_size_bytes = parse_size(max_size) if max_size else None
    space_keys_set = load_list_from_arg(space_keys)
    exclude_space_keys_set = load_list_from_arg(exclude_space_keys)
    try:
        mod_after = parse_date(modified_after) if modified_after else None
    except ValueError as e:
        logging.error(e)
        mod_after = None
    try:
        mod_before = parse_date(modified_before) if modified_before else None
    except ValueError as e:
        logging.error(e)
        mod_before = None
    
    # Parse created_in_year - support single year or comma-separated list
    created_years = None
    if created_in_year:
        try:
            years_str = [y.strip() for y in created_in_year.split(',')]
            created_years = [int(y) for y in years_str if y]
        except ValueError:
            logging.error(f"Invalid year format in --created-in-year: {created_in_year}")
            created_years = None

    # Parse modified_in_year - support single year or comma-separated list
    modified_years = None
    if modified_in_year:
        try:
            years_str = [y.strip() for y in modified_in_year.split(',')]
            modified_years = [int(y) for y in years_str if y]
        except ValueError:
            logging.error(f"Invalid year format in --modified-in-year: {modified_in_year}")
            modified_years = None

    if not keywords and not patterns:
        logging.error("No search criteria provided. Specify --keywords, --regex, or --regex-file")
        return

    if keywords:
        logging.info(f"Loaded keywords: {len(keywords)}")
    if patterns:
        logging.info(f"Loaded regex patterns: {len(patterns)}")
    
    if public_only:
        logging.info("Searching for public spaces...")
        spaces = get_public_spaces(base_url, headers)
    else:
        spaces = get_all_spaces(base_url, auth, headers)
    
    if space_keys_set:
        spaces = [s for s in spaces if s.get("key", "").upper() in {k.upper() for k in space_keys_set}]
    spaces = [s for s in spaces if s.get("key", "").upper() not in {k.upper() for k in exclude_space_keys_set}]
    spaces.sort(key=lambda s: s.get("key", ""))
    if resume_from:
        spaces = [s for s in spaces if s.get("key", "").upper() >= resume_from.strip().upper()]
    
    logging.info(f"Found/filtered spaces: {len(spaces)}")

    # cloud_id is no longer needed: downloads use the 302-redirect mechanism
    # (REST endpoint → api.media.atlassian.com with pre-signed token).
    cloud_id = None

    findings_set = set() if no_duplicates else None
    
    total_secrets = 0
    total_files = 0
    total_pages = 0
    
    # Debug mode: stop after 5 findings
    debug_limit = 5 if debug_mode else None
    
    # Store all findings for author alerts
    all_findings_data = []
    
    with open(output_file, "w", encoding="utf-8", newline="") as csvfile:
        writer = csv.writer(csvfile)
        if include_attachments:
            writer.writerow([
                "Space Name", "Page Title", "File Title", "File Extension", "File URL", 
                "Page URL", "Matched Keyword", "Finding Type", "Matched Value", "Last Editor Email"
            ])
            # Dummy row to absorb the first-row bug
            writer.writerow([
                "DUMMY_ROW_DELETE_ME", "DUMMY", "DUMMY", "txt", "http://dummy.com/file", 
                "http://dummy.com/page", "DUMMY_KEYWORD", "dummy", "DUMMY_VALUE", "dummy@dummy.com"
            ])
        else:
            writer.writerow([
                "Space Name", "Page Title", "Page URL", "Matched Keyword", 
                "Finding Type", "Matched Value", "Last Editor Email"
            ])
            # Dummy row to absorb the first-row bug
            writer.writerow([
                "DUMMY_ROW_DELETE_ME", "DUMMY", "http://dummy.com/page", "DUMMY_KEYWORD", 
                "dummy", "DUMMY_VALUE", "dummy@dummy.com"
            ])
        
        logging.info("🔧 Added dummy row after headers to prevent first-row bug")

        for space in spaces:
            secrets_in_space, files_in_space, pages_in_space = process_space(
                base_url, auth, headers, space, keywords, patterns, writer, csvfile,
                include_attachments, allowed_types, excluded_types, max_size_bytes, 
                mod_after, mod_before, created_years, modified_years, no_duplicates, 
                secret_max_length, scan_images_only, archive_support, findings_set, 
                debug_limit, total_secrets, cloud_id=cloud_id
            )
            total_secrets += secrets_in_space
            total_files += files_in_space
            total_pages += pages_in_space
            
            # Debug mode: stop if reached limit
            if debug_mode and total_secrets >= debug_limit:
                logging.warning(f"🔍 DEBUG MODE: Reached {debug_limit} findings limit. Stopping scan and generating report...")
                break

    end_time = time.time()
    duration = end_time - start_time
    duration_str = f"{int(duration // 60)}m {int(duration % 60)}s"
    
    # Debug mode summary
    if debug_mode:
        logging.warning("=" * 80)
        logging.warning(f"🔍 DEBUG MODE: Stopped after {total_secrets} findings (limit: {debug_limit})")
        logging.warning("=" * 80)
    
    global_summary = f"[GLOBAL SUMMARY] Total scan duration: {duration_str}, {total_pages} pages scanned, {total_secrets} secrets found"
    if include_attachments:
        global_summary += f", {total_files} files scanned"
    logging.info(global_summary)
    
    logging.info(f"Results saved in: {output_file}")
    
    # Determine base name for output files from -o flag (strip extension if provided)
    output_base = re.sub(r'\.(csv|xlsx|json)$', '', output_file, flags=re.IGNORECASE)

    # Create XLSX report (name derived from -o base)
    output_dir = os.path.dirname(os.path.abspath(output_file))
    if include_attachments:
        # In recursive call for files mode, output_file is already e.g. results_files.csv
        # so we just use output_base + ".xlsx" to avoid double "_files_files"
        xlsx_file = output_base + ".xlsx"
    else:
        xlsx_file = output_base + ".xlsx"

    xlsx_created = False
    if create_xlsx_report(output_file, xlsx_file):
        logging.info(f"XLSX report created: {xlsx_file}")
        xlsx_created = True
    else:
        logging.warning("XLSX report creation skipped or failed")

    # Create JSON report if --json flag was set
    if export_json:
        json_file = output_base + ("_files.json" if include_attachments and mode == "files" else ".json")
        scan_stats = {
            "spaces_scanned": len(spaces),
            "pages_scanned": total_pages,
            "files_scanned": total_files if include_attachments else 0,
            "secrets_found": total_secrets,
            "duration": duration_str,
        }
        export_findings_to_json(output_file, json_file, include_attachments, scan_stats)
    
    # Send email if configured
    if email_sender and email_recipient:
        if not aws_region:
            aws_region = "eu-central-1"  # default
        
        logging.info("Preparing to send email report...")
        
        # Count unique affected spaces and pages from CSV
        affected_spaces_count, affected_pages_count = count_affected_from_csv(output_file)
        
        # Generate email body
        email_body = generate_email_summary(
            total_secrets=total_secrets,
            total_pages=total_pages,
            total_files=total_files,
            spaces_scanned=len(spaces),
            duration_str=duration_str,
            include_attachments=include_attachments,
            affected_spaces=affected_spaces_count,
            affected_pages=affected_pages_count
        )
        
        # Prepare subject with scan mode indicator
        mode_label = ""
        if mode == "files":
            mode_label = " (Files)"
        elif mode == "pages":
            mode_label = " (Pages)"
        else:
            mode_label = " (General)"
        
        if total_secrets > 0:
            subject = f"CRITICAL: Confluence Secrets Scanner{mode_label} - {total_secrets} Secrets Found"
        else:
            subject = f"Confluence Secrets Scanner{mode_label} - No Secrets Found"
        
        # Send email with attachment
        attachment = xlsx_file if xlsx_created and os.path.exists(xlsx_file) else None
        
        send_email_with_attachment(
            subject=subject,
            body_text=email_body,
            sender=email_sender,
            recipient=email_recipient,
            aws_region=aws_region,
            attachment_path=attachment
        )
    else:
        logging.info("Email notification disabled (no sender/recipient configured)")
    
    # Send individual author alerts if enabled
    if alert and total_secrets > 0:
        if not email_sender:
            logging.warning("\n⚠️  To send alerts, --email-sender must be specified")
        else:
            logging.info("\n📨 Sending individual alerts to page editors...")
            
            # Parse CSV to group findings by editor email
            editors_findings = {}
            
            try:
                with open(output_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    header = next(reader)  # Skip header
                    
                    # Skip dummy row if present
                    first_row = next(reader, None)
                    if first_row and "DUMMY_ROW_DELETE_ME" in first_row[0]:
                        pass  # Already skipped
                    else:
                        # Process first row if it's not dummy
                        if first_row:
                            f.seek(0)  # Reset
                            next(reader)  # Skip header again
                    
                    for row in reader:
                        if not row or "DUMMY_ROW_DELETE_ME" in row[0]:
                            continue
                        
                        if include_attachments:
                            # With attachments: Space Name, Page Title, File Title, File Extension, File URL, Page URL, Keyword, Type, Value, Email
                            if len(row) >= 10:
                                editor_email = row[9].strip()
                                editor_name = editor_email  # Use email as name since we don't have display name
                                
                                finding = {
                                    'space_name': row[0],
                                    'page_title': row[1],
                                    'file_title': row[2],
                                    'file_extension': row[3],
                                    'file_url': row[4],
                                    'page_url': row[5],
                                    'keyword': row[6],
                                    'finding_type': row[7],
                                    'matched_value': row[8],
                                    'editor_email': editor_email,
                                    'is_attachment': True
                                }
                        else:
                            # Without attachments: Space Name, Page Title, Page URL, Keyword, Type, Value, Email
                            if len(row) >= 7:
                                editor_email = row[6].strip()
                                editor_name = editor_email  # Use email as name
                                
                                finding = {
                                    'space_name': row[0],
                                    'page_title': row[1],
                                    'page_url': row[2],
                                    'keyword': row[3],
                                    'finding_type': row[4],
                                    'matched_value': row[5],
                                    'editor_email': editor_email,
                                    'is_attachment': False
                                }
                        
                        # Skip if no valid email
                        if not editor_email or editor_email == 'N/A' or '@' not in editor_email or 'unknown' in editor_email:
                            continue
                        
                        if editor_email not in editors_findings:
                            editors_findings[editor_email] = {
                                'name': editor_name,
                                'findings': []
                            }
                        editors_findings[editor_email]['findings'].append(finding)
                
                # Send alerts to each editor
                logging.info(f"   Found {len(editors_findings)} editor(s) with secrets")
                
                success_count = 0
                for editor_email, data in editors_findings.items():
                    editor_name = data['name']
                    editor_findings = data['findings']
                    
                    logging.info(f"   📧 {editor_name} ({editor_email}): {len(editor_findings)} secret(s)")
                    
                    if send_author_alert(
                        author_email=editor_email,
                        author_name=editor_name,
                        author_findings=editor_findings,
                        aws_region=aws_region,
                        sender_email=email_sender,
                        security_contact=security_contact,
                        security_wiki=security_wiki,
                        has_attachments=include_attachments
                    ):
                        success_count += 1
                
                logging.info(f"\n   ✅ Successfully sent: {success_count}/{len(editors_findings)}")
                
            except Exception as e:
                logging.error(f"Error processing alerts: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Improved Confluence Scanner v5 - Scan for secrets and keywords with enhanced regex support and XLSX output",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic scan (pages only, default)
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt

  # Scan only files/attachments
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    -m files --filetype docx,pdf,json

  # Scan both pages and files separately (creates 2 reports)
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    -m both --filetype docx,pdf,json

  # Public-only scan
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --public-only --regex-file regex.txt
    
  # Scan with author alerts
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    --email-sender security@company.com --alert --security-contact appsec@company.com
    
  # Send results to multiple recipients
  python3 confluence.py --base-url https://your-org.atlassian.net \\
    --username user@example.com --token YOUR_TOKEN --regex-file regex.txt \\
    --email-sender security@company.com --email-recipient "appsec@company.com,team@company.com"
        """
    )
    # ── 1. Connection ────────────────────────────────────────────────────────────
    conn = parser.add_argument_group("Connection")
    conn.add_argument("--base-url",    help="Base Confluence URL")
    conn.add_argument("--username",    help="API login (email) — not required with --public-only")
    conn.add_argument("--token",       help="API token — not required with --public-only")
    conn.add_argument("--public-only", action="store_true",
                      help="Scan only public spaces without authentication")

    # ── 2. Scan mode ─────────────────────────────────────────────────────────────
    mode_grp = parser.add_argument_group("Scan mode")
    mode_grp.add_argument("-m", "--mode", choices=["files", "both"],
                          help="'files' — attachments only; 'both' — pages + files (2 reports). "
                               "Default: pages only")
    mode_grp.add_argument("--space-keys",         help="Space keys to scan (comma-separated or file path)")
    mode_grp.add_argument("--exclude-space-keys", help="Space keys to exclude (comma-separated or file path)")
    mode_grp.add_argument("--resume-from",        help="Resume scanning from this space key")
    mode_grp.add_argument("--no-duplicates",      action="store_true", help="Exclude duplicate findings")

    # ── 3. Patterns ───────────────────────────────────────────────────────────────
    pat = parser.add_argument_group("Patterns (at least one required)")
    pat.add_argument("--keywords",    help="File with keywords, one per line")
    pat.add_argument("--regex-file",  help="Regex file in 'Name:::Regex:::GroupIndex' format")
    pat.add_argument("--regex",       help="Single regex pattern (legacy)")
    pat.add_argument("--trufflehog-patterns", "-tp", dest="trufflehog_yaml",
                     nargs="?", const="trufflehog.yaml", default=None,
                     metavar="FILE",
                     help="TruffleHog YAML file with detectors (default: `trufflehog.yaml` in current directory)")
    pat.add_argument("--trufflehog-keywords", "-tk", dest="trufflehog_keywords",
                     metavar="KEYWORDS",
                     help="Include only TruffleHog detectors whose 'keywords' match any of these. "
                          "Example: -tk aws,api,internal")
    pat.add_argument("--trufflehog-exclude-keywords", "-tek", dest="trufflehog_exclude_keywords",
                     metavar="KEYWORDS",
                     help="Exclude TruffleHog detectors whose 'keywords' match any of these. "
                          "Mutually exclusive with -tk. Example: -tek gateway,arn")

    # ── 4. File / attachment options ──────────────────────────────────────────────
    files = parser.add_argument_group("File / attachment options  (require -m files or -m both)")
    files.add_argument("--filetype",         help="File types to scan, e.g. docx,pdf,json")
    files.add_argument("--exclude-filetype", help="File types to exclude, e.g. pdf")
    files.add_argument("--max-size",         help="Max file size to analyse, e.g. 2mb, 500kb")
    files.add_argument("--scan-images-only", action="store_true", help="Scan only images via OCR")
    files.add_argument("--archive-support",  action="store_true",
                       help="Unpack and scan archives (zip, tar, etc.)")

    # ── 5. Date filters ───────────────────────────────────────────────────────────
    dates = parser.add_argument_group("Date filters")
    dates.add_argument("--modified-after",  help="Pages modified after this date (D.M.Y or D/M/Y)")
    dates.add_argument("--modified-before", help="Pages modified before this date (D.M.Y or D/M/Y)")
    dates.add_argument("--created-in-year", help="Filter by creation year (not last modified). Single year or comma-separated list, e.g. 2025 or 2024,2025")
    dates.add_argument("--modified-in-year", help="Filter by last-modified year (not creation date). Single year or comma-separated list, e.g. 2026 or 2025,2026")

    # ── 6. Output ─────────────────────────────────────────────────────────────────
    out = parser.add_argument_group("Output")
    out.add_argument("-o", "--output", default="confluence_results.csv",
                     help="Output base name (extensions .csv / .xlsx / .json added automatically)")
    out.add_argument("--json", action="store_true",
                     help="Export findings to JSON in addition to XLSX")
    out.add_argument("--secret-max-length", type=int, default=None,
                     help="Max characters shown in 'Matched Value' column. Default: full value")

    # ── 7. Email notifications ────────────────────────────────────────────────────
    email = parser.add_argument_group("Email notifications  (requires AWS SES)")
    email.add_argument("--email-sender",    help="Verified SES sender address")
    email.add_argument("--email-recipient", help="Recipient(s) for scan report, comma-separated")
    email.add_argument("--aws-region",      default="eu-central-1",
                       help="AWS SES region (default: eu-central-1)")

    # ── 8. Author alerts ──────────────────────────────────────────────────────────
    alerts = parser.add_argument_group("Author alerts  (notify page editors directly)")
    alerts.add_argument("--alert", action="store_true",
                        help="Send individual alerts to editors who leaked secrets")
    alerts.add_argument("--security-contact", default="security@company.com",
                        help="Security team email shown in alert (default: security@company.com)")
    alerts.add_argument("--security-wiki",
                        help="Security docs URL included in author alerts (optional)")

    # ── 9. Misc ───────────────────────────────────────────────────────────────────
    misc = parser.add_argument_group("Misc")
    misc.add_argument("--config", help="JSON config file — values are overridden by CLI flags")
    misc.add_argument("--debug",  action="store_true",
                      help="Stop after first 5 findings and generate report")
    
    args = parser.parse_args()
    
    validation_errors = validate_arguments(args)
    if validation_errors:
        setup_logging()
        logging.error("Argument validation errors:")
        for error in validation_errors:
            logging.error(f"  - {error}")
        logging.info("Use --help for more information.")
        exit(1)

    main(
        base_url=args.base_url,
        username=args.username,
        token=args.token,
        keywords_file=args.keywords,
        regex_file=args.regex_file,
        single_regex=args.regex,
        output_file=args.output,
        filetypes=args.filetype,
        exclude_filetypes=args.exclude_filetype,
        max_size=args.max_size,
        public_only=args.public_only,
        space_keys=args.space_keys,
        exclude_space_keys=args.exclude_space_keys,
        modified_after=args.modified_after,
        modified_before=args.modified_before,
        created_in_year=args.created_in_year,
        modified_in_year=getattr(args, 'modified_in_year', None),
        no_duplicates=args.no_duplicates,
        resume_from=args.resume_from,
        scan_images_only=args.scan_images_only,
        archive_support=args.archive_support,
        config=args.config,
        email_sender=args.email_sender,
        email_recipient=args.email_recipient,
        aws_region=args.aws_region,
        secret_max_length=args.secret_max_length,
        debug_mode=args.debug,
        alert=args.alert,
        security_contact=args.security_contact,
        security_wiki=args.security_wiki,
        mode=getattr(args, 'mode', None),
        trufflehog_yaml=args.trufflehog_yaml,
        trufflehog_keywords=getattr(args, 'trufflehog_keywords', None),
        trufflehog_exclude_keywords=getattr(args, 'trufflehog_exclude_keywords', None),
        export_json=args.json,
    )
